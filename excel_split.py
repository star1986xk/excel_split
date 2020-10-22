from openpyxl import *
import shutil
import re
from settings import split_path

def split(filename):
    wb = load_workbook(filename)
    sheetnames = wb.sheetnames
    for n, name in enumerate(sheetnames[2:]):
        # 拷贝一个excel并打开编辑
        shutil.copy(filename, split_path + name + '.xlsx')
        n_wb = load_workbook(split_path + name + '.xlsx')

        n_sheetnames = n_wb.sheetnames

        # 修改封面
        n_ws = n_wb[n_sheetnames[0]]
        n_ws.cell(row=14, column=4).value = 'OPPO ' + name

        # 删除column
        n_ws = n_wb[n_sheetnames[1]]
        for c in reversed(range(3, n_ws.max_column + 1)):
            if c <= (n * 7) + 2 or c > (n * 7) + 9:
                n_ws.delete_cols(c)

        # 替换公式中字母
        column_count = n_ws.max_column
        for r in n_ws.rows:
            for c in range(0, column_count):
                if isinstance(r[c].value, str) and '=' in r[c].value:
                    result = re.findall('([A-Z]+)[0-9]+', r[c].value)
                    for li in result:
                        if li != 'SUM':
                            num = 0
                            for char_n, i in enumerate(reversed(li)):
                                num += (ord(i) - 64) * (26 ** char_n)
                            r[c].value = str(r[c].value).replace(li, chr(num - n * 7 + 64))

        # 删除sheet
        for n_name in n_sheetnames[2:]:
            if name != n_name:
                del n_wb[n_name]

        n_wb.save(split_path + name + '.xlsx')
