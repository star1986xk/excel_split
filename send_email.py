from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText  # 邮件正文
from email.mime.multipart import MIMEMultipart
from email.header import Header  # 邮件头
from email.mime.base import MIMEBase  # MIME子类的基类
from email import encoders  # 导入编码器
from settings import address_book


def send(host, port, user, password):
    # 加载excel文件
    print('正在加载文件......')
    wb = load_workbook(address_book, data_only=True)
    sheet1 = wb.active

    # 登陆邮箱
    smtp_obj = smtplib.SMTP_SSL(host, port=port)  # 邮箱发送服务器
    smtp_obj.login(user, password)  # 邮箱用户名，密码(授权码)
    count = 0
    table_col_html = '<thead>'  # 表头
    for row in sheet1.iter_rows(min_row=1):
        count += 1
        if count == 1:
            for col in row:
                table_col_html += f"<th>{col.value}<t/th>"
            table_col_html += '</thead>'
            continue
        else:
            row_test = '<tr>'  # 开始一行
            for cell in row:
                # print(cell.value,end=',')
                row_test += f"<td>{cell.value}</td>"
            row_test += "</tr>"  # 结束一行
            datatime = row[3]
            name = row[0]
            city = row[2]
            number = row[4]
            staff_email = row[1].value
            # print(staff_email,name)
            mail_body_context = f"""
              <h3>{name.value},您好：</h3>
              以下是 {city.value}{datatime.value} 分品牌激活数据,请查收！
          """
            # 邮件正文是MIMEText
            msgtext = MIMEText(mail_body_context, 'html', 'utf-8')
            # 邮件对象
            num = number.value
            msg = MIMEMultipart('related')
            msg.attach(msgtext)
            msg['From'] = user  # 发送者
            msg['To'] = staff_email
            msg['cc'] = 'luxiaotianoppo@163.com'
            msg['Subject'] = Header('W31各品牌激活数据汇总', 'utf-8').encode()  # 主题
            # 附件文件定义
            # 创建一个MIMEText对象，附加表格文件（week.xlsx）
            filename = './数据/' + city.value + '.xlsx'
            attachfile = MIMEBase('applocation', 'octet-stream')  # 创建对象指定主要类型和次要类型
            attachfile.set_payload(open(filename, 'rb').read())  # 将消息内容设置为有效载荷
            attachfile.add_header('Content-Disposition', 'attachment',
                                  filename=('utf-8', '', city.value + '.xlsx'))  # 扩展标题设置
            encoders.encode_base64(attachfile)
            msg.attach(attachfile)  # 附加对象加入到msg
            # 发邮
            smtp_obj.sendmail(user, [staff_email, 'luxiaotianoppo@163.com'], msg.as_string())
            print(f"（{city.value}）数据成功发送到（{name.value}）（{staff_email}）邮箱中.....")
            if num % 10 == 0:
                smtp_obj.quit()
                smtp_obj = smtplib.SMTP_SSL(host, port=port)  # 邮箱发送服务器
                smtp_obj.login(user, password)  # 邮箱用户名，密码(授权码)
                print('请稍等，重新连接中....')
    smtp_obj.quit()
    print('程序运行成功')
